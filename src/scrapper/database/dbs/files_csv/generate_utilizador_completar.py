from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from collections import defaultdict
import copy
import datetime

DB_PATH = './database_export.xlsx'
TEMPLATE_PATH = './Utilizador_Completar.xlsx'
OUTPUT_PATH = './Utilizador_Completar_novo.xlsx'

#bloat
COURSE_MAP = {
    'LEIC': 'L.EIC',
    'MEIC': 'M.EIC',
    'MESW': 'MESW',
    'MIA':  'M.IA',
    'MM':   'MM',
}
# acronyms L.EIC','M.EIC','MESW','M.IA','MM'

def load_db_data(db_path):
    wb = load_workbook(db_path, data_only=True)

    # courses: db_acronym -> id
    courses = {}
    for row in wb['course'].iter_rows(min_row=2, values_only=True):
        courses[row[3]] = row[0]

    # course_units: course_id -> [{'id','name','acronym','semester'}]
    cus_by_course = defaultdict(list)
    for row in wb['course_unit'].iter_rows(min_row=2, values_only=True):
        cus_by_course[row[1]].append({'id': row[0], 'name': row[2], 'acronym': row[3], 'semester': row[5]})

    # course_metadata: course_unit_id -> course_unit_year
    cu_year = {}
    for row in wb['course_metadata'].iter_rows(min_row=2, values_only=True):
        # course_id, course_unit_id, course_unit_year, ects
        cu_id = row[1]
        year = row[2]
        if cu_id not in cu_year:  # guard against duplicates
            cu_year[cu_id] = year

    # student overlap
    student_to_cus = defaultdict(set)
    for row in wb['students_in_course_units'].iter_rows(min_row=2, values_only=True):
        student_to_cus[row[0]].add(row[1])

    
    overlap = defaultdict(int)
    for cu_ids in student_to_cus.values():
        cu_list = list(cu_ids)
        for i in range(len(cu_list)):
            for j in range(len(cu_list)):
                overlap[(cu_list[i], cu_list[j])] += 1

    return courses, cus_by_course, overlap, cu_year

def copy_sheet_from_template(wb_src, wb_dst, sheet_name):
    ws_src = wb_src[sheet_name]
    ws_dst = wb_dst.create_sheet(title=sheet_name)

    for row in ws_src.iter_rows():
        for cell in row:
            new_cell = ws_dst.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = copy.copy(cell.font)
                new_cell.fill = copy.copy(cell.fill)
                new_cell.border = copy.copy(cell.border)
                new_cell.alignment = copy.copy(cell.alignment)
                new_cell.number_format = cell.number_format

    for col_letter, col_dim in ws_src.column_dimensions.items():
        ws_dst.column_dimensions[col_letter].width = col_dim.width

    for row_dim_key, row_dim in ws_src.row_dimensions.items():
        ws_dst.row_dimensions[row_dim_key].height = row_dim.height

    for dv in ws_src.data_validations.dataValidation:
        ws_dst.add_data_validation(copy.copy(dv))

    for merge in ws_src.merged_cells.ranges:
        ws_dst.merge_cells(str(merge))

    return ws_dst

def build_uc_sheet(ws, course_units, cu_year):
    arial = Font(name='Arial')

    # Headers row 1
    headers = [None, 'Ultima avaliação por exame final', 'Data da Ultima avaliação', 'ANO']
    for col, val in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=val)
        cell.font = arial

    # Data validation for col B (boolean)
    dv_bool = DataValidation(type='list', formula1='"TRUE,FALSE"', allow_blank=True, sqref=f'B2:B{len(course_units)+1}')
    ws.add_data_validation(dv_bool)

    # Data validation for col C (date)
    dv_date = DataValidation(type='date', operator='greaterThanOrEqual', formula1='DATE(2000,1,1)',
                             allow_blank=True, showErrorMessage=True,
                             prompt='Data inválida - Introduza uma data válida (ex: 01/01/2026)',
                             sqref=f'C2:C{len(course_units)+1}')
    ws.add_data_validation(dv_date)

    for i, cu in enumerate(course_units, start=2):
        ws.cell(row=i, column=1, value=cu['acronym']).font = arial
        ws.cell(row=i, column=2, value=False).font = arial
        ws.cell(row=i, column=3, value=None).font = arial
        ws.cell(row=i, column=3).number_format = 'DD/MM/YYYY'
        year_val = cu_year.get(cu['id'], None)
        ws.cell(row=i, column=4, value=year_val).font = arial

    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 10

def build_inscritos_sheet(ws, course_units, overlap):
    arial = Font(name='Arial')
    header_fill = PatternFill('solid', fgColor='4472C4')
    header_font = Font(name='Arial', bold=True, color='FFFFFF')
    diag_fill = PatternFill('solid', fgColor='D9D9D9')
    center = Alignment(horizontal='center', vertical='center')

    # Top-left corner
    corner = ws.cell(row=1, column=1, value='UC \\ UC')
    corner.fill = header_fill
    corner.font = header_font
    corner.alignment = center

    cu_ids = [cu['id'] for cu in course_units]
    cu_names = [cu['acronym'] or cu['name'][:15] for cu in course_units]

    # Column headers
    for col_idx, name in enumerate(cu_names, start=2):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    # Row headers + matrix
    for row_idx, (cu_id_row, name_row) in enumerate(zip(cu_ids, cu_names), start=2):
        cell = ws.cell(row=row_idx, column=1, value=name_row)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

        for col_idx, cu_id_col in enumerate(cu_ids, start=2):
            value = overlap.get((cu_id_row, cu_id_col), 0)
            c = ws.cell(row=row_idx, column=col_idx, value=value)
            c.alignment = center
            c.font = arial
            if cu_id_row == cu_id_col:
                c.fill = diag_fill

    ws.column_dimensions['A'].width = 18
    for col_idx in range(2, len(cu_ids) + 2):
        ws.column_dimensions[get_column_letter(col_idx)].width = 10

def main():
    print("Loading database...")
    courses, cus_by_course, overlap, cu_year = load_db_data(DB_PATH)

    print("Loading template...")
    wb_orig = load_workbook(TEMPLATE_PATH)

    wb = Workbook()
    wb.remove(wb.active)

    # Copy Critérios and Calendário sheets from template exactly
    for sheet_name in ['Critérios', 'Calendário']:
        print(f"Copying sheet: {sheet_name}")
        copy_sheet_from_template(wb_orig, wb, sheet_name)

    # Build Unidades_Curriculares and Inscritos_UCs for each course
    for display_name, db_acronym in COURSE_MAP.items():
        if db_acronym not in courses:
            print(f"  Skipping {display_name} - not in database")
            continue

        course_id = courses[db_acronym]
        all_units = cus_by_course.get(course_id, [])

        for semester in [1, 2]:
            units = [cu for cu in all_units if cu['semester'] == semester]
            if not units:
                continue

            sheet_suffix = f"{display_name}_S{semester}"

            # Unidades_Curriculares sheet
            uc_sheet_name = f"UCs_{sheet_suffix}"
            print(f"  Building sheet: {uc_sheet_name} ({len(units)} units)")
            ws_uc = wb.create_sheet(title=uc_sheet_name)
            build_uc_sheet(ws_uc, units, cu_year)

            # Inscritos_UCs sheet
            ins_sheet_name = f"Inscritos_{sheet_suffix}"
            print(f"  Building sheet: {ins_sheet_name}")
            ws_ins = wb.create_sheet(title=ins_sheet_name)
            build_inscritos_sheet(ws_ins, units, overlap)

    wb.save(OUTPUT_PATH)
    print(f"\nSaved: {OUTPUT_PATH}")

if __name__ == '__main__':
    main()
