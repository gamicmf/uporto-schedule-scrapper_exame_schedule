import sqlite3
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# ── Config ──────────────────────────────────────────────
DB_PATH = "../database.db"
OUTPUT_FILE = "./database_export.xlsx"
# ────────────────────────────────────────────────────────

def export_to_excel(db_path, output_file):
    con = sqlite3.connect(db_path)
    cur = con.cursor()

    # Get all table names
    cur.execute("SELECT name FROM sqlite_master WHERE type='table'")
    tables = [row[0] for row in cur.fetchall()]

    print(f"Found {len(tables)} tables: {tables}\n")

    wb = Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    for table in tables:
        cur.execute(f"PRAGMA table_info({table})")
        columns = [row[1] for row in cur.fetchall()]
        cur.execute(f"SELECT * FROM {table}")
        rows = cur.fetchall()

        ws = wb.create_sheet(title=table)
        ws.append(columns)
        for row in rows:
            ws.append(list(row))

        print(f"✓ Sheet '{table}' → {len(rows)} rows, {len(columns)} columns")

    # ── Overlap matrix sheet ─────────────────────────────
    add_overlap_sheet(wb, con)

    wb.save(output_file)
    con.close()
    print(f"\nExcel file saved to: {os.path.abspath(output_file)}")


def add_overlap_sheet(wb, con):
    cur = con.cursor()

    # Get all course units with their names
    cur.execute("""
        SELECT cu.id, cu.acronym, cu.name
        FROM course_unit cu
        ORDER BY cu.id
    """)
    course_units = cur.fetchall()

    if not course_units:
        print("⚠ No course units found, skipping overlap sheet")
        return

    cu_ids = [row[0] for row in course_units]
    cu_labels = [f"{row[1]}" for row in course_units]  # use acronym as label

    # Build overlap map: (cu_id_1, cu_id_2) -> shared student count
    cur.execute("""
        SELECT 
            a.course_unit_id AS cu1,
            b.course_unit_id AS cu2,
            COUNT(*) AS shared
        FROM students_in_course_units a
        JOIN students_in_course_units b
            ON a.student_id = b.student_id
        GROUP BY a.course_unit_id, b.course_unit_id
    """)
    overlap_data = {(row[0], row[1]): row[2] for row in cur.fetchall()}

    ws = wb.create_sheet(title="Student Overlap")

    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")
    diag_fill = PatternFill("solid", fgColor="D9D9D9")
    center = Alignment(horizontal="center", vertical="center")

    # Write column headers (row 1)
    ws.cell(row=1, column=1, value="UC \\ UC")
    ws.cell(row=1, column=1).fill = header_fill
    ws.cell(row=1, column=1).font = header_font

    for col_idx, label in enumerate(cu_labels, start=2):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    # Write row headers and matrix values
    for row_idx, (cu_id_row, label) in enumerate(zip(cu_ids, cu_labels), start=2):
        # Row header
        cell = ws.cell(row=row_idx, column=1, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

        for col_idx, cu_id_col in enumerate(cu_ids, start=2):
            value = overlap_data.get((cu_id_row, cu_id_col), 0)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = center
            # Diagonal (same UC) gets grey
            if cu_id_row == cu_id_col:
                cell.fill = diag_fill

    # Auto-size first column
    ws.column_dimensions["A"].width = 20
    for col_idx in range(2, len(cu_ids) + 2):
        ws.column_dimensions[get_column_letter(col_idx)].width = 10

    print(f"✓ Sheet 'Student Overlap' → {len(cu_ids)}x{len(cu_ids)} matrix")


if __name__ == "__main__":
    export_to_excel(DB_PATH, OUTPUT_FILE)